from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

# 打开 Excel 文件
wb = load_workbook('details/paper_citation_details.xlsx')
ws = wb.active

# 读取每行的数据
papers_info = []
paper_dict = {}  # 存储论文名字和编号的映射关系
for row in ws.iter_rows(values_only=True):
    paper_id = row[0]
    paper_title = row[1]
    references = set(row[2:])  # 使用集合去除重复的参考文献
    papers_info.append({'编号': paper_id, '名字': paper_title, '参考文献': references})
    paper_dict[paper_title] = paper_id

# 筛选有效的参考文献
valid_references = set()
for paper_info in papers_info:
    paper_title = paper_info['名字']
    references = paper_info['参考文献']
    valid_references.update(references & set(paper_dict.keys()))  # 取交集，即有效的参考文献

# 更新有效参考文献的编号
ref_to_id = {ref: paper_dict[ref] for ref in valid_references}

# 更新论文的参考文献
valid_papers_info = []
for paper_info in papers_info:
    paper_id = paper_info['编号']
    references = paper_info['参考文献']
    valid_references_in_paper = references & valid_references
    if valid_references_in_paper:  # 仅保留存在有效参考文献的论文
        paper_info['参考文献'] = [ref_to_id[ref] for ref in valid_references_in_paper]
        valid_papers_info.append(paper_info)

# 创建新的 Excel 文件并写入数据
wb_new = Workbook()
ws_new = wb_new.active
ws_new.append(['论文编号', '参考文献编号'])

for paper_info in valid_papers_info:
    paper_id = paper_info['编号']
    references = sorted(paper_info['参考文献'])  # 按编号排序
    ws_new.append([paper_id] + references)

# 生成时间戳作为文件名
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

# 保存新的 Excel 文件到项目文件的xlsx文件包里
file_name = f'mapping/论文引用关系_{timestamp}.xlsx'
wb_new.save(file_name)
