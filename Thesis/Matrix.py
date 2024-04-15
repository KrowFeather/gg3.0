import pandas as pd
from openpyxl import load_workbook


def build_paper_matrix():
    # 读取生成的论文引用关系 Excel 文件
    wb = load_workbook('./Thesis/mapping/mapping.xlsx')
    ws = wb.active
    print(wb.sheetnames)
    # 提取所有不重复节点
    nodes = set()
    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        # 跳过第一行提示信息
        if row_num == 0:
            continue
        paper_id = row[0]
        references = row[1:]
        nodes.add(paper_id)
        for ref_id in references:
            if ref_id:
                nodes.add(ref_id)

    # 补充缺失的节点
    missing_nodes = set(range(1, 51)) - nodes
    nodes = sorted(nodes.union(missing_nodes))

    # 创建空的邻接矩阵
    adj_matrix = pd.DataFrame(0, index=nodes, columns=nodes)

    # 遍历每行的数据，更新邻接矩阵的值
    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        # 跳过第一行提示信息
        if row_num == 0:
            continue
        paper_id = row[0]
        references = row[1:]
        for ref_id in references:
            if ref_id:
                adj_matrix.loc[paper_id, ref_id] = 1

    # 创建新的一行和一列，填充为0
    new_row = pd.Series(0, index=adj_matrix.columns, name=0)
    new_column = pd.Series(0, index=adj_matrix.index, name=0)

    # 插入新的一行作为第一行
    adj_matrix = new_row.to_frame().T._append(adj_matrix, ignore_index=True)
    # 插入新的一列作为第一列
    adj_matrix.insert(0, 0, 0)

    final_matrix = [[0 for _ in range(len(adj_matrix))] for _ in range(len(adj_matrix))]

    for i in range(0, len(final_matrix)):
        for j in range(0, len(final_matrix)):
            final_matrix[i][j] = adj_matrix.loc[i][j]

    return final_matrix
