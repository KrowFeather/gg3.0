from openpyxl import load_workbook
import csv
from datetime import datetime
import os

# 打开 Excel 文件
wb = load_workbook('details/paper_citation_details.xlsx')
ws = wb.active

# 读取每行的数据
papers_info = []
paper_dict = {}  # 存储论文名字和编号的映射关系
for row in ws.iter_rows(values_only=True):
    paper_id = row[0]
    paper_title = row[1]
    references = set(row[2:])  # 使用集合去除重复的参考文献
    papers_info.append({'编号': paper_id, '名字': paper_title, '参考文献': references})
    paper_dict[paper_title] = paper_id

# 筛选有效的参考文献
valid_references = set()
for paper_info in papers_info:
    paper_title = paper_info['名字']
    references = paper_info['参考文献']
    valid_references.update(references & set(paper_dict.keys()))  # 取交集，即有效的参考文献

# 更新有效参考文献的编号
ref_to_id = {ref: paper_dict[ref] for ref in valid_references}

# 更新论文的参考文献
valid_papers_info = []
for paper_info in papers_info:
    paper_id = paper_info['编号']
    references = paper_info['参考文献']
    valid_references_in_paper = references & valid_references
    if valid_references_in_paper:  # 仅保留存在有效参考文献的论文
        paper_info['参考文献'] = [ref_to_id[ref] for ref in valid_references_in_paper]
        valid_papers_info.append(paper_info)

# 生成时间戳作为文件名
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

# 保存到CSV文件
csv_file_name = f'csv/论文引用关系_{timestamp}.csv'
with open(csv_file_name, 'w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    writer.writerow(["论文编号", "论文名字", "参考文献编号", "参考文献名字", "引用"])
    for paper_info in valid_papers_info:
        paper_id = paper_info['编号']
        paper_title = paper_info['名字']
        for ref_id in paper_info['参考文献']:
            ref_name = next((key for key, val in paper_dict.items() if val == ref_id), None)
            writer.writerow([paper_id, paper_title, ref_id, ref_name, "引用"])
print(f"CSV文件已创建成功: {csv_file_name}")
