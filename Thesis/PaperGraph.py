import networkx as nx
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def generate_paper_graph():
    # 读取生成的论文引用关系 Excel 文件
    wb = load_workbook('./Thesis/mapping/mapping.xlsx')
    ws = wb.active

    # 创建有向图
    G = nx.DiGraph()

    # 遍历每行的数据，添加边
    for row in ws.iter_rows(values_only=True):
        if row[0] == '论文编号':  # 跳过标题行
            continue
        paper_id = row[0]
        references = row[1:]
        for ref_id in references:
            if ref_id:
                G.add_edge(paper_id, ref_id)

    # 绘制有向图
    plt.figure(figsize=(10, 8))  # 设置画布大小
    pos = nx.spring_layout(G)  # 使用 Spring 布局算法
    nx.draw(G, pos, with_labels=True, node_size=300, node_color="skyblue", font_size=12, arrowsize=15)
    plt.title("Paper Citation Network")  # 设置标题
    plt.show()

    return G
